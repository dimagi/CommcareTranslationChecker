from __future__ import print_function
import sys
import os
import datetime
import argparse
import traceback as tb
import openpyxl as xl


##### DEFINE ERRORS #####
class CellError(Exception):
    '''
    Exception that occurs while processing a given cell.
    '''
    def __init__(self, cell, e, verbose=False):
        '''
        CellError initializer.

        Input:
        cell(xl.cell.cell.Cell): Cell in which the error occurred
        e(Exception): Exception that occurred
        verbose(bool [opt]): Boolean indicating whether verbose flag is on
        '''
        self.cell = cell 
        self.e = e
        self.verbose = verbose

    def __str__(self):
        '''
        Print exception and cell it occurred on.
        '''
        if self.verbose:
            tb.print_exc(e)
        return "Exception %s occurred in Sheet %s Cell %s" % (str(e),self.cell.parent.title, self.cell.coordinate)

class WorksheetError(Exception):
    '''
    Exception that occurs while processing a given worksheet.
    '''
    def __init__(self, ws, e, verbose=False):
        '''
        WorksheetError initializer.

        Input:
        ws(xl.worksheet.worksheet.Worksheet): Worksheet in which the error occurred
        e(Exception): Exception that occurred
        verbose(bool [opt]): Boolean indicating whether verbose flag is on
        '''
        self.ws = ws 
        self.e = e
        self.verbose = verbose

    def __str__(self):
        '''
        Print exception and cell it occurred on.
        '''
        if self.verbose:
            tb.print_exc(e)
        print("Exception %s occurred in Sheet %s" % (str(e),self.title))


##### DEFINE METHODS #####

def parseArguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="Location of Translation file to check", type=str, required = True)
    parser.add_argument("--columns", help="[Opt] Comma-separated list of column names to check. By default, all columns that start with 'default_' will be checked.", type=str, default=None)
    parser.add_argument("--base-column", help="[Opt] Name of column that others are to be compared against. Warnings are flagged for all columns that do not match the base-column. Defaults to leftmost column in columns.", type=str, default=None, dest='baseColumn')
    parser.add_argument("--ignore-order", help="[Opt] If passed, the order in which output value tags appear will not be considered when comparing cells against each other. This is useful if the order of the output value tags is different between columns because of differences in word orders between the languages involved.", action="store_true", default=False, dest='ignoreOrder')
    parser.add_argument("-v", "--verbose",  help="[Opt] If passed, output will be printed to the screen pointing out which rows of the file have issues.", action="store_true", default = False)
    parser.add_argument("--output-folder", help = "[Opt] Folder in which any output files should be passed. Defaults to 'commcareTranslationChecker_Output' folder relative to folder from which the script is called. Can be relative or absolute path.", type=str, default = "commcareTranslationChecker_Output", dest='outputFolder')
    parser.add_argument("--no-output-file", help = "[Opt] If passed, no output file will be created.", action="store_false", default = True, dest = "createOutputFileFlag")
    parser.add_argument("--configuration-sheet", help = "[Opt] Specify which sheet contains configuration information about modules and forms. Defaults to 'Modules_and_forms'", type=str, default = "Modules_and_forms", dest='configurationSheet')
    parser.add_argument("--configuration-sheet-column", help = "[Opt] specify which column in the configuration sheet contains expected sheet names. Defaults to 'sheet_name'", type=str, default = "sheet_name", dest='configurationSheetColumnName')
    parser.add_argument("--debug-mode", "-d", action="store_true", default=False, dest="debugMode")
    return parser.parse_args()

def convertCellToOutputValueList(cell):
    '''
    Convert an Excel cell to a list of <output value...> tags contained within the cell

    Input:
    cell (xl.cell.cell.Cell): Cell whose contents are to be parsed

    Output:
    List of unicode objects, each representing an instance of <output value...> in cell. Any of these values that appear suspicious will be prepended with "ILL-FORMATTED TAG : "
    '''
    openTag = "<output value=\""
    closeTag ="\"/>"
    outputList = []
    currentIndex = 0
    try:
        while cell.value[currentIndex:].find(openTag) != -1:
            currentIndex += cell.value[currentIndex:].find(openTag) + len(openTag)
            closeTagIndex = cell.value[currentIndex:].find(closeTag)
            if closeTagIndex != -1:
                outputValue = cell.value[currentIndex:closeTagIndex + currentIndex]
                if outputValue.find(openTag) == -1:
                    outputList.append(outputValue)
                else:
                    outputList.append("ILL-FORMATTED TAG : " + outputValue)
            else:
                outputValue = cell.value[currentIndex:]
                print("closeTag not found for " + outputValue)
                outputList.append("ILL-FORMATTED TAG : " + outputValue)
    except TypeError as e:
        print(e) 
        return []

    return outputList

def createOutputCell(cell, wsOut):
    '''
    Make a copy of a Cell object into the exact same coordinates in the target Worksheet.

    Input:
    cell (xl.cell.cell.Cell): Cell whose contents and coordinates are to be copied
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet to which the cell's contents are to be copied

    Output:
    New Cell in wsOut
    '''
    try:
        newCell = wsOut.cell(coordinate = cell.coordinate) 
        newCell.value = cell.value
        newCell.style = xl.styles.Style(alignment = xl.styles.Alignment(wrap_text = True))
        return newCell
    except Exception, e:
        print("FATAL ERROR creating output cell for worksheet %s cell %s (writing to output worksheet %s) : %s" % (cell.parent.title, cell.coordinate, wsOut.title, str(e)))
        exit(-1)

def getOutputCell(cell, wsOut):
    '''
    Fetch an existing Cell object from wsOut corresponding to the coordinates of cell.

    Input:
    cell (xl.cell.cell.Cell): Cell whose coordinates are to be used to pinpoint target cell 
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet from which corresponding cell is to be pulled

    Output:
    Cell objects from wsOut corresponding to coordinates of cell 
    '''

    return wsOut[cell.coordinate]


def checkRowForMismatch(row, columnDict, baseColumnIdx = None, ignoreOrder = False, wsOut = None, mismatchFlagIdx = None):
    '''
    Check all of the given columns in a row provided for any mismatch in the columns' OutputValueList 

    Input:
    row(list): list of openyxl.cell.cell.Cell objects representing a single row in an Excel sheet 
    columnDict(dict): dictionary mapping column index to column name, representing every column to be checked against the baseColumn 
    baseColumnIdx(int [opt]): Index of the column to be considered "correct." Defaults to lowest-indexed column in columnDict.
    ignoreOrder(bool [opt]): If True, the order in which output values appear will be ignored for purposes of comparing cells. Otherwise, the order will matter. Defaults to False.
    wsOut(xl.worksheet.worksheet.Worksheet [opt]): Worksheet whose corresponding cell should be filled with Red if a mismatch occurs. Defaults to None.
    mismatchFlagIdx(int [opt]): Column index where the mismatchFlag value should be printed in wsOut

    Output:
    Tuple consisting of a single-element dictionary mapping the baseColumn's index to its outputValueList, and a dictionary mapping the column indexes of mismatched cells to a tuple consisting of the associated cell's OutputValueList and a list of mismatchTypes. wsOut altered so that every Cell that is mismatched is filled with Red, and mismatchFlag column filled with "Y" if there was a mismatch in the row, "N" otherwise.
    '''
    mismatchDict = {}
    baseColumnDict=  {}

    baseOutputValueList = None

    mismatchFillStyle = xl.styles.Style(fill = xl.styles.PatternFill(fgColor = xl.styles.colors.Color(xl.styles.colors.RED), fill_type = "solid"), alignment = xl.styles.Alignment(wrap_text = True))

    ## Get columnDictKeyList for Python3
    columnDictKeyList = list(columnDict.keys())

    ## Build baseColumnDict
    if baseColumnIdx is None:
        baseColumnIdx = sorted(columnDictKeyList)[0]
    baseOutputValueList = convertCellToOutputValueList(row[baseColumnIdx])
    if ignoreOrder:
        baseOutputValueList = sorted(baseOutputValueList)
    baseColumnDict = {baseColumnIdx : baseOutputValueList}

    for colIdx in columnDictKeyList:
        try:
            curOutputValueList = convertCellToOutputValueList(row[colIdx])
            if ignoreOrder:
                curOutputValueList = sorted(curOutputValueList)
            if colIdx != baseColumnIdx and baseOutputValueList != curOutputValueList:
                ## Determine how everything is mismatched
                mismatchTypes = []

                ## Determine whether any ill-formatted tags exist:
                illFormattedValueList = []
                for value in curOutputValueList:
                    if value.startswith("ILL-FORMATTED TAG : "):
                        illFormattedValueList.append(value[20:])
                if illFormattedValueList != []:
                    mismatchTypes.append("Ill-Formatted Tags - " + ",".join(illFormattedValueList)) 

                ## Determine whether any values missing from current list
                missingValueList = []
                for value in baseOutputValueList:
                    if value not in curOutputValueList:
                        missingValueList.append(value)
                if missingValueList != []:
                    mismatchTypes.append("Missing Values - " + ",".join(missingValueList))

                ## Determine whether extra values have been added in current list
                extraValueList = []
                for value in curOutputValueList:
                    if value not in baseOutputValueList:
                        extraValueList.append(value)
                if extraValueList != []:
                    mismatchTypes.append("Extra Values - " + ",".join(extraValueList))

                ## Determine if, after considering missing/extra values, there are sort issues
                if not ignoreOrder:
                    baseListIndex = 0
                    for value in curOutputValueList:
                        if value not in extraValueList:
                            while baseOutputValueList[baseListIndex] in missingValueList:
                                baseListIndex += 1
                            if value != baseOutputValueList[baseListIndex]:
                                mismatchTypes.append("Out of Order")
                                break 
                            baseListIndex += 1

                mismatchDict[colIdx] = (curOutputValueList, mismatchTypes)
                if wsOut:
                    cellOut = getOutputCell(row[colIdx], wsOut)
                    cellOut.style = mismatchFillStyle

                print(mismatchTypes)
        except AttributeError as e:
            print(e)

    mismatchCell =wsOut.cell(row = getOutputCell(row[0], wsOut).row, column = 1).offset(column = mismatchFlagIdx)
    if len(mismatchDict) > 0:
        mismatchCell.value = "Y"
        mismatchCell.style = mismatchFillStyle
    else:
        mismatchCell.value = "N"

    return (baseColumnDict, mismatchDict)


def checkConfigurationSheet(wb, ws, configurationSheetColumnName, wsOut, verbose = False):
    '''
    Check that the workbook contains one sheet for every corresponding entry in the configurationSheetColumn of ws, and highlight all cells in wsOut that represent sheets that don't exist.

    Input:
    wb (xl.workbook.workbook.Workbook): Workbook containing sheets to check against configurationSheetColumn
    ws (xl.worksheet.worksheet.Worksheet): Worksheet containing the column to check 
    configurationSheetColumnName (str): Name of column to compare sheet names against
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet to print highlighted cells to 
    verbose (boolen [opt]): If passed, prints each missing sheet to the screen

    Output:
    List of sheets that are missing from the Workbook. If configurationSheetColumnName does not exist in ws, returns None
    '''
    mismatchFillStyle = xl.styles.Style(fill = xl.styles.PatternFill(fgColor = xl.styles.colors.Color(xl.styles.colors.RED), fill_type = "solid"), alignment = xl.styles.Alignment(wrap_text = True))

    ## Check that the configuration column exists at all
    colIdx = None
    for headerIdx, cell in enumerate(ws.rows[0]):
        if cell.value == configurationSheetColumnName:
            colIdx = headerIdx
    if colIdx == None:
        print(configurationSheetColumnName, " not found in ", ws.title, ". Skipping sheet check.")
        return None

    ## Iterate over configuration column, flagging red if corresponding sheet does not exist
    for cell in ws.columns[colIdx][1:]:
        if cell.value not in (sheet.title for sheet in wb):
            getOutputCell(cell, wsOut).style = mismatchFillStyle
            if verbose:
                print("WARNING %s: This sheet is missing from the workbook: " % (cell.value,))


def main(argv):
    args = parseArguments()
    try:
        wb = xl.load_workbook(args.file)
        if args.verbose:
            print("Workbook Loaded")
    except xl.exceptions.InvalidFileException as e:
        print("Invalid File: %s" % (str(e),))
        if args.debugMode:
            tb.print_exc(e)
        exit(-1)

    ## Open new Workbook
    wbOut = xl.Workbook()
    wbOut.remove_sheet(wbOut.active)

    ## Summary lists
    wsMismatchDict = {}

    ## Iterate through WorkSheets
    for ws in wb:
        try:
            wbOut.create_sheet(title = ws.title)
            wsOut = wbOut[ws.title]


            ## Dictionary mapping column index to column name
            defaultColumnDict = {}

            maxHeaderIdx = 0
            ## Find all columns of format "default_[CODE]"
            for headerIdx, cell in enumerate(ws.rows[0]):
                ## First, copy cell into new workbook
                cellOut = createOutputCell(cell, wsOut)
                if args.columns:
                    if cell.value in args.columns:
                        defaultColumnDict[headerIdx] = cell.value
                elif cell.value and cell.value[:8] == "default_":
                    defaultColumnDict[headerIdx] = cell.value
                if headerIdx > maxHeaderIdx:
                    maxHeaderIdx = headerIdx
            ## If defaultColumnDict is empty, skip processing
            ## Otherwise, reate header cell in wsOut for mismatchFlag
            if len(defaultColumnDict) != 0:
                mismatchFlagIdx = maxHeaderIdx + 1
                wsOut.cell("A1").offset(column = mismatchFlagIdx).value = "mismatchFlag"


                for rowIdx, row in enumerate(ws.rows[1:]):
                    ## First, copy every cell into new workbook
                    for cell in row:
                        cellOut = createOutputCell(cell, wsOut)

                    ## Fetch baseColumn information
                    baseColumnIdx = None
                    if args.baseColumn:
                        for colIdx in defaultColumnDict.keys():
                            if defaultColumnDict[colIdx] == args.baseColumn:
                                baseColumnIdx = colIdx 

                    ## Check row for mismatch and print results
                    rowCheckResults = checkRowForMismatch(row, defaultColumnDict, baseColumnIdx, args.ignoreOrder, wsOut, mismatchFlagIdx)
                    if len(rowCheckResults[1]) > 0:
                        if ws.title not in wsMismatchDict.keys():
                            wsMismatchDict[ws.title] = 1
                        else:
                            wsMismatchDict[ws.title] += 1
                        if args.verbose:
                            baseColumnName = defaultColumnDict[list(rowCheckResults[0].keys())[0]]
                            mismatchColumnNames = ",".join(defaultColumnDict[i] for i in rowCheckResults[1].keys())
                            print("WARNING %s row %s: the output values in %s do not match %s" % (ws.title, rowIdx+2, mismatchColumnNames, baseColumnName))
            elif args.verbose:
                print("WARNING %s: No columns found for comparison" % (ws.title,))

            ## If ws is a configuration sheet, run the configuration check
            if ws.title == args.configurationSheet:
                checkConfigurationSheet(wb, ws, args.configurationSheetColumnName, wsOut, args.verbose)
        except Exception, e:
            print("FATAL ERROR in worksheet %s : %s" % (ws.title, str(e)))
            if args.debugMode:
                tb.print_exc(e)
            exit(-1)

    ## Save workbook and print summary
    if len(wsMismatchDict) > 0:
        if args.createOutputFileFlag:
            tsString = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            fileBasename = os.path.splitext(os.path.basename(args.file))[0]
            outputFolder = args.outputFolder
            outputFileName = os.path.join(outputFolder,"%s_%s_Output.xlsx" % (fileBasename, tsString))
            ## Create the output directory if it does not exist
            if not os.path.exists(os.path.dirname(outputFileName)):
                try:
                    os.makedirs(os.path.dirname(outputFileName),)
                    print("Output directory did not exist, created %s" % (os.path.dirname(outputFileName),))
                except OSError as e:
                    if e.errorno != e.EEXIST:
                        print("ERROR CREATING OUTPUT DIRECTORY : %s" % (str(e),))
                        if args.debugMode:
                            tb.print_exc(e)
            wbOut.save(outputFileName)
            print("There were issues with the following worksheets, see %s for details:" % (outputFileName,))
        else:
            print("There were issues with the following worksheets:")
        for key in wsMismatchDict.keys():
            print("%s : %s row%s mismatched" % (key, wsMismatchDict[key], "" if wsMismatchDict[key]==1 else "s"))

def entryPoint():
    main(sys.argv[1:])

if __name__ == "__main__":
    entryPoint()