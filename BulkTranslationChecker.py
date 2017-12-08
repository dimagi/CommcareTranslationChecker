import sys
import re
import argparse
import openpyxl as xl

def parseArguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="Location of Translation file to check", type=str)
    parser.add_argument("--columnList", help="[Opt] Comma-separated list of column names to check. By default, all columns that start with 'default_' will be checked.", type=str, default=None)
    parser.add_argument("--baseColumn", help="[Opt] Name of column that others are to be compared against. Warnings are flagged for all columns that do not match the baseColumn. Defaults to leftmost column in columnList.", type=str, default=None)
    parser.add_argument("--ignoreOrder", help="[Opt] If passed, the order in which output value tags appear will not be considered when comparing cells against each other. This is useful if the order of the output value tags is different between columns because of differences in word orders between the languages involved.", action="store_true", default=False)
    parser.add_argument("-v", "--verbose",  help="[Opt] If passed, output will be printed to the screen pointing out which rows of the file have issues.", action="store_true", default = False)
    return parser.parse_args()

def convertCellToOutputValueList(cell):
    '''
    Convert an Excel cell to a list of <output value...> tags contained within the cell

    Input:
    cell (xl.cell.cell.Cell): Cell whose contents are to be parsed

    Output:
    List of unicode objects, each representing an instance of <output value...> in cell
    '''
    #### First pass: find an instance of "<output value" and pull whole string until next ">"
    #### Second pass: for each "<" after output value, ignore the next ">"
    openTag = "output value=\""
    closeTag ="\"/>"
    outputList = []
    currentIndex = 0
    try:
        while cell.value[currentIndex:].find(openTag) != -1:
            currentIndex += cell.value[currentIndex:].find(openTag) + len(openTag)
            outputList.append(cell.value[currentIndex:cell.value[currentIndex:].find(closeTag) + currentIndex])
    except TypeError, e:
        return []

    return outputList

def createOutputCell(cell, wsOut):
    '''
    Make a copy of a Cell object into the exact same coordinates in the target Worksheet.

    Input:
    cell (xl.cell.cell.Cell): Cell whose contents and coordinates are to be copied
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet to which the cell's contents are to be copied

    Output:
    New Cell in wsOut
    '''
    newCell = wsOut.cell(coordinate = cell.coordinate) 
    newCell.value = cell.value
    newCell.style = xl.styles.Style(alignment = xl.styles.Alignment(wrap_text = True))
    return newCell

def getOutputCell(cell, wsOut):
    '''
    Fetch an existing Cell object from wsOut corresponding to the coordinates of cell.

    Input:
    cell (xl.cell.cell.Cell): Cell whose coordinates are to be used to pinpoint target cell 
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet from which corresponding cell is to be pulled

    Output:
    Cell objects from wsOut corresponding to coordinates of cell 
    '''

    return wsOut[cell.coordinate]


def checkRowForMismatch(row, columnDict, baseColumnIdx = None, ignoreOrder = False, wsOut = None, mismatchFlagIdx = None):
    '''
    Check all of the given columns in a row provided for any mismatch in the columns' OutputValueList 

    Input:
    row(list): list of openyxl.cell.cell.Cell objects representing a single row in an Excel sheet 
    columnDict(dict): dictionary mapping column index to column name, representing every column to be checked against the baseColumn 
    baseColumnIdx(int [opt]): Index of the column to be considered "correct." Defaults to lowest-indexed column in columnDict.
    ignoreOrder(bool [opt]): If True, the order in which output values appear will be ignored for purposes of comparing cells. Otherwise, the order will matter. Defaults to False.
    wsOut(xl.worksheet.worksheet.Worksheet [opt]): Worksheet whose corresponding cell should be filled with Red if a mismatch occurs. Defaults to None.
    mismatcFlagIdx(int [opt]): Column index where the mismatchFlag value should be printed in wsOut

    Output:
    Tuple consisting of a single-element dictionary mapping the baseColumn's index to its outputValueList, and a dictionary mapping the column indexes of mismatched cells to their OutputValueList. wsOut altered so that every Cell that is mismatched is filled with Red, and mismatchFlag column filled with "Y" if there was a mismatch in the row, "N" otherwise.
    '''
    mismatchDict = {}
    baseColumnDict=  {}

    baseOutputValueList = None

    mismatchFillStyle = xl.styles.Style(fill = xl.styles.PatternFill(fgColor = xl.styles.colors.Color(xl.styles.colors.RED), fill_type = "solid"), alignment = xl.styles.Alignment(wrap_text = True))

    ## Build baseColumnDict
    if baseColumnIdx is None:
        baseColumnIdx = sorted(columnDict.keys())[0]
    baseOutputValueList = convertCellToOutputValueList(row[baseColumnIdx])
    if ignoreOrder:
        baseOutputValueList = sorted(baseOutputValueList)
    baseColumnDict = {baseColumnIdx : baseOutputValueList}

    for colIdx in columnDict.keys():
        try:
            curOutputValueList = convertCellToOutputValueList(row[colIdx])
            if ignoreOrder:
                curOutputValueList = sorted(curOutputValueList)
            if colIdx != baseColumnIdx and baseOutputValueList != curOutputValueList:
                mismatchDict[colIdx] = curOutputValueList
                if wsOut:
                    cellOut = getOutputCell(row[colIdx], wsOut)
                    cellOut.style = mismatchFillStyle
        except AttributeError, e:
            pass

    mismatchCell =wsOut.cell(row = getOutputCell(row[0], wsOut).row, column = 1).offset(column = mismatchFlagIdx)
    if len(mismatchDict) > 0:
        mismatchCell.value = "Y"
        mismatchCell.style = mismatchFillStyle
    else:
        mismatchCell.value = "N"

    return (baseColumnDict, mismatchDict)


def main(argv):
    args = parseArguments()
    try:
        wb = xl.load_workbook(args.file)
        print "Workbook Loaded"
    except xl.exceptions.InvalidFileException, e:
        print "Invalid File!"
        exit(-1)

    ## Open new Workbook
    wbOut = xl.Workbook()
    wbOut.remove_sheet(wbOut.active)

    ## Iterate through WorkSheets
    for ws in wb:
        wbOut.create_sheet(title = ws.title)
        wsOut = wbOut[ws.title]

        ## Dictionary mapping column index to column name
        defaultColumnDict = {}

        maxHeaderIdx = 0
        ## Find all columns of format "default_[CODE]"
        for headerIdx, cell in enumerate(ws.rows[0]):
            ## First, copy cell into new workbook
            cellOut = createOutputCell(cell, wsOut)
            if args.columnList:
                if cell.value in args.columnList:
                    defaultColumnDict[headerIdx] = cell.value
            elif cell.value[:8] == "default_":
                defaultColumnDict[headerIdx] = cell.value
            if headerIdx > maxHeaderIdx:
                maxHeaderIdx = headerIdx
        ## Create header cell in wsOut for mismatchFlag
        mismatchFlagIdx = maxHeaderIdx + 1
        wsOut.cell("A1").offset(column = mismatchFlagIdx).value = "mismatchFlag"


        for rowIdx, row in enumerate(ws.rows[1:]):
            ## First, copy every cell into new workbook
            for cell in row:
                cellOut = createOutputCell(cell, wsOut)

            ## Fetch baseColumn information
            baseColumnIdx = None
            if args.baseColumn:
                for colIdx in defaultColumnDict.keys():
                    if defaultColumnDict[colIdx] == args.baseColumn:
                        baseColumnIdx = colIdx 

            ## Check row for mismatch and print results
            rowCheckResults = checkRowForMismatch(row, defaultColumnDict, baseColumnIdx, args.ignoreOrder, wsOut, mismatchFlagIdx)
            if len(rowCheckResults[1]) > 0:
                baseColumnName = defaultColumnDict[rowCheckResults[0].keys()[0]]
                baseColumnOutputValueList = rowCheckResults[0][rowCheckResults[0].keys()[0]]
                mismatchColumnNames = ",".join(defaultColumnDict[i] for i in rowCheckResults[1].keys())
                print "WARNING %s row %s: the output values in %s do not match %s" % (ws.title, rowIdx+2, mismatchColumnNames, baseColumnName)
                # print "%s: %s \n%s: %s" % (curOutputValueList, curCellValue.encode('utf8'), baseOutputValueList, baseCellValue.encode('utf8'))

    ## Save workbook
    wbOut.save("results.xlsx")


if __name__ == "__main__":
    main(sys.argv[1:])